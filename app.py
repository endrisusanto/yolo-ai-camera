import cv2
import math
import time
import os
import datetime
import pytz
import sqlite3
from flask import Flask, render_template, Response, request, jsonify, send_from_directory, send_file
from ultralytics import YOLO
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
import face_recognition

app = Flask(__name__)

# Configuration
CAPTURE_FOLDER = 'static/captures'
FACES_FOLDER = 'faces'
DATABASE_FILE = 'visionguard.db'
os.makedirs(CAPTURE_FOLDER, exist_ok=True)
os.makedirs(FACES_FOLDER, exist_ok=True)

# Initialize Database
def init_db():
    conn = sqlite3.connect(DATABASE_FILE)
    cursor = conn.cursor()
    
    # Phone alerts table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS phone_alerts (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            filename TEXT NOT NULL,
            timestamp TEXT NOT NULL,
            date TEXT NOT NULL,
            type TEXT NOT NULL,
            description TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    # Sitting sessions table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sitting_sessions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            person_id INTEGER NOT NULL,
            duration INTEGER NOT NULL,
            timestamp TEXT NOT NULL,
            date TEXT NOT NULL,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    
    conn.commit()
    conn.close()

# Initialize database on startup
init_db()

# Timezone GMT+7
TZ = pytz.timezone('Asia/Jakarta')

# Load Models
model_seg_n = YOLO('yolov8n-seg.pt')
model_pose_n = YOLO('yolov8n-pose.pt')

# Face Recognition - Load known faces
known_face_encodings = []
known_face_names = []

def load_known_faces():
    """Load face encodings from faces folder"""
    global known_face_encodings, known_face_names
    
    known_face_encodings = []
    known_face_names = []
    
    if not os.path.exists(FACES_FOLDER):
        print(f"Faces folder not found: {FACES_FOLDER}")
        return
    
    for filename in os.listdir(FACES_FOLDER):
        if filename.lower().endswith(('.jpg', '.jpeg', '.png')):
            filepath = os.path.join(FACES_FOLDER, filename)
            try:
                # Load image
                image = face_recognition.load_image_file(filepath)
                # Get face encoding
                encodings = face_recognition.face_encodings(image)
                
                if len(encodings) > 0:
                    # Use first face found
                    known_face_encodings.append(encodings[0])
                    # Use filename without extension as name
                    name = os.path.splitext(filename)[0]
                    known_face_names.append(name)
                    print(f"Loaded face: {name}")
                else:
                    print(f"No face found in: {filename}")
            except Exception as e:
                print(f"Error loading {filename}: {e}")
    
    print(f"Total faces loaded: {len(known_face_names)}")

# Load known faces on startup
load_known_faces()

# Global state
settings = {
    "fps_limit": 60,
    "mode": "fast",
    "conf_threshold": 0.25,
    "low_light_mode": False
}

current_fps = 0
captures = []
captures = []
person_sessions = {}
sitting_history = []
last_capture_time = 0
CAPTURE_COOLDOWN = 3.0
SITTING_PERSIST_TIME = 1.0
next_person_id = 0
recognized_faces = {}  # Store recognized faces with their names



camera = cv2.VideoCapture(0)

def apply_low_light_enhancement(frame):
    """Enhance image for low light conditions using CLAHE"""
    # Convert to LAB color space
    lab = cv2.cvtColor(frame, cv2.COLOR_BGR2LAB)
    l, a, b = cv2.split(lab)
    
    # Apply CLAHE (Contrast Limited Adaptive Histogram Equalization) to L-channel
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
    cl = clahe.apply(l)
    
    # Merge and convert back to BGR
    limg = cv2.merge((cl,a,b))
    enhanced = cv2.cvtColor(limg, cv2.COLOR_LAB2BGR)
    return enhanced


def get_local_time():
    """Get current time in GMT+7"""
    return datetime.datetime.now(TZ)

def calculate_angle(a, b, c):
    """Calculate angle between three points."""
    a = np.array(a)
    b = np.array(b)
    c = np.array(c)
    
    radians = np.arctan2(c[1]-b[1], c[0]-b[0]) - np.arctan2(a[1]-b[1], a[0]-b[0])
    angle = np.abs(radians*180.0/np.pi)
    
    if angle > 180.0:
        angle = 360 - angle
        
    return angle

def calculate_iou(box1, box2):
    """Calculate Intersection over Union between two bounding boxes"""
    x1_min, y1_min, x1_max, y1_max = box1
    x2_min, y2_min, x2_max, y2_max = box2
    
    inter_x_min = max(x1_min, x2_min)
    inter_y_min = max(y1_min, y2_min)
    inter_x_max = min(x1_max, x2_max)
    inter_y_max = min(y1_max, y2_max)
    
    if inter_x_max < inter_x_min or inter_y_max < inter_y_min:
        return 0.0
    
    inter_area = (inter_x_max - inter_x_min) * (inter_y_max - inter_y_min)
    
    box1_area = (x1_max - x1_min) * (y1_max - y1_min)
    box2_area = (x2_max - x2_min) * (y2_max - y2_min)
    union_area = box1_area + box2_area - inter_area
    
    if union_area == 0:
        return 0.0
    
    return inter_area / union_area

def get_person_bbox(keypoints):
    """Get bounding box from keypoints"""
    valid_points = keypoints[keypoints[:, 0] > 0]
    if len(valid_points) == 0:
        return None
    
    x_min = valid_points[:, 0].min()
    y_min = valid_points[:, 1].min()
    x_max = valid_points[:, 0].max()
    y_max = valid_points[:, 1].max()
    
    return [x_min, y_min, x_max, y_max]

def match_person_to_session(bbox, current_time, active_ids):
    """Match current person to existing session using IoU and Distance"""
    global next_person_id
    
    best_match_id = None
    best_iou = 0.0
    min_dist = float('inf')
    
    # Calculate centroid of current person
    cx = (bbox[0] + bbox[2]) / 2
    cy = (bbox[1] + bbox[3]) / 2
    
    # First pass: IoU
    for person_id, session in person_sessions.items():
        if person_id in active_ids:
            continue
            
        if 'bbox' in session:
            iou = calculate_iou(bbox, session['bbox'])
            if iou > best_iou and iou > 0.3:
                best_iou = iou
                best_match_id = person_id
    
    # Second pass: Distance (if no IoU match)
    if best_match_id is None:
        for person_id, session in person_sessions.items():
            if person_id in active_ids:
                continue
                
            if 'bbox' in session:
                s_bbox = session['bbox']
                s_cx = (s_bbox[0] + s_bbox[2]) / 2
                s_cy = (s_bbox[1] + s_bbox[3]) / 2
                
                dist = math.sqrt((cx - s_cx)**2 + (cy - s_cy)**2)
                
                # Match if within 200 pixels
                if dist < 200 and dist < min_dist:
                    min_dist = dist
                    best_match_id = person_id

    if best_match_id is None:
        best_match_id = next_person_id
        next_person_id += 1
    
    return best_match_id

def check_head_down(keypoints):
    """Check if head is looking down"""
    if len(keypoints) > 6:
        nose = keypoints[0][:2]
        left_eye = keypoints[1][:2]
        right_eye = keypoints[2][:2]
        left_shoulder = keypoints[5][:2]
        right_shoulder = keypoints[6][:2]
        
        if all([nose[0] > 0, left_eye[0] > 0, right_eye[0] > 0, 
                left_shoulder[0] > 0, right_shoulder[0] > 0]):
            
            eye_y = (left_eye[1] + right_eye[1]) / 2
            shoulder_y = (left_shoulder[1] + right_shoulder[1]) / 2
            nose_to_eye_dist = nose[1] - eye_y
            eye_to_shoulder_dist = shoulder_y - eye_y
            
            if eye_to_shoulder_dist > 0:
                ratio = nose_to_eye_dist / eye_to_shoulder_dist
                if ratio > 0.3:
                    return True
    
    return False

def check_sitting_advanced(keypoints, chairs_boxes):
    """Advanced sitting detection"""
    if len(keypoints) < 17:
        return "Standing", 0
    
    left_hip = keypoints[11][:2]
    right_hip = keypoints[12][:2]
    left_knee = keypoints[13][:2]
    right_knee = keypoints[14][:2]
    left_ankle = keypoints[15][:2]
    right_ankle = keypoints[16][:2]
    left_shoulder = keypoints[5][:2]
    right_shoulder = keypoints[6][:2]
    
    hip_x = (left_hip[0] + right_hip[0]) / 2
    hip_y = (left_hip[1] + right_hip[1]) / 2
    
    sitting_score = 0
    knee_angle = 0
    
    if all([left_hip[0] > 0, left_knee[0] > 0, left_ankle[0] > 0]):
        left_angle = calculate_angle(left_hip, left_knee, left_ankle)
        if 60 < left_angle < 120:
            sitting_score += 1
            knee_angle = left_angle
    
    if all([right_hip[0] > 0, right_knee[0] > 0, right_ankle[0] > 0]):
        right_angle = calculate_angle(right_hip, right_knee, right_ankle)
        if 60 < right_angle < 120:
            sitting_score += 1
            if knee_angle == 0:
                knee_angle = right_angle
    
    if left_knee[0] > 0 and right_knee[0] > 0:
        avg_knee_y = (left_knee[1] + right_knee[1]) / 2
        hip_knee_diff = hip_y - avg_knee_y
        
        if -50 < hip_knee_diff < 100:
            sitting_score += 1
    
    if all([left_shoulder[0] > 0, right_shoulder[0] > 0, hip_x > 0]):
        shoulder_y = (left_shoulder[1] + right_shoulder[1]) / 2
        torso_vertical_dist = hip_y - shoulder_y
        shoulder_x = (left_shoulder[0] + right_shoulder[0]) / 2
        torso_horizontal_dist = abs(hip_x - shoulder_x)
        
        if torso_vertical_dist > 0 and torso_horizontal_dist < torso_vertical_dist * 0.5:
            sitting_score += 1
    
    if sitting_score >= 2:
        # If pose strongly suggests sitting, we don't strictly need a chair
        return "Sitting", knee_angle
    
    return "Standing", 0

def draw_timer_card(frame, x, y, timer_text, status="SITTING"):
    """Draw glassmorphism timer card"""
    padding = 20
    text_size = cv2.getTextSize(timer_text, cv2.FONT_HERSHEY_SIMPLEX, 0.9, 2)[0]
    label_size = cv2.getTextSize(status, cv2.FONT_HERSHEY_SIMPLEX, 0.5, 1)[0]
    
    card_width = max(text_size[0], label_size[0]) + padding * 2
    card_height = 70
    
    card_x = x - card_width // 2
    card_y = y - 100
    
    card_x = max(10, min(card_x, frame.shape[1] - card_width - 10))
    card_y = max(10, card_y)
    
    overlay = frame.copy()
    
    cv2.rectangle(overlay, 
                 (card_x, card_y), 
                 (card_x + card_width, card_y + card_height),
                 (20, 20, 20), -1)
    
    cv2.rectangle(overlay, 
                 (card_x, card_y), 
                 (card_x + card_width, card_y + card_height),
                 (100, 100, 100), 2)
    
    alpha = 0.75
    cv2.addWeighted(overlay, alpha, frame, 1 - alpha, 0, frame)
    
    # Color based on status
    # Orange for Sitting, Blue for Standing
    color = (255, 149, 0) if status == "SITTING" else (0, 149, 255)
    
    cv2.rectangle(frame,
                 (card_x, card_y),
                 (card_x + card_width, card_y + 4),
                 color, -1)
    
    label_x = card_x + (card_width - label_size[0]) // 2
    cv2.putText(frame, status, 
               (label_x, card_y + 25),
               cv2.FONT_HERSHEY_SIMPLEX, 0.5, (150, 150, 150), 1)
    
    timer_x = card_x + (card_width - text_size[0]) // 2
    cv2.putText(frame, timer_text,
               (timer_x, card_y + 55),
               cv2.FONT_HERSHEY_SIMPLEX, 0.9, (255, 255, 255), 2)

def generate_frames():
    global current_fps, last_capture_time, person_sessions, sitting_history
    
    prev_time = time.time()
    
    while True:
        loop_start = time.time()
        
        success, frame = camera.read()
        if not success:
            break
            
        time_diff = time.time() - prev_time
        target_interval = 1.0 / settings["fps_limit"]
        if time_diff < target_interval:
            time.sleep(target_interval - time_diff)
            
        prev_time = time.time()
        
        # Apply low light enhancement if enabled
        if settings.get("low_light_mode", False):
            frame = apply_low_light_enhancement(frame)
        
        conf = settings["conf_threshold"]
        
        # Optimized YOLO inference - smaller resolution (416 instead of 640)
        # Use half precision if available for faster inference
        results_seg = model_seg_n(frame, conf=conf, verbose=False, imgsz=416, half=True)
        results_pose = model_pose_n(frame, conf=conf, verbose=False, imgsz=416, half=True)
        
        # Faster plotting with reduced line width
        annotated_frame = results_seg[0].plot(line_width=1, font_size=0.5)
        pose_overlay = results_pose[0].plot(line_width=1, font_size=0.5)
        annotated_frame = cv2.addWeighted(annotated_frame, 0.7, pose_overlay, 0.3, 0)
        
        chairs_boxes = []
        phone_detected = False
        
        # Expanded detection details
        detected_objects = []
        
        if results_seg[0].boxes is not None:
            for box in results_seg[0].boxes:
                cls_id = int(box.cls[0])
                xyxy = box.xyxy[0].cpu().numpy()
                conf_score = float(box.conf[0])
                
                # 56: Chair, 67: Cell phone, 63: Laptop, 41: Cup, 73: Book
                if cls_id == 56:
                    chairs_boxes.append(xyxy)
                elif cls_id == 67:
                    phone_detected = True
                    detected_objects.append("Phone")
                elif cls_id == 63:
                    detected_objects.append("Laptop")
                elif cls_id == 41:
                    detected_objects.append("Cup")
                elif cls_id == 73:
                    detected_objects.append("Book")
        
        # Check for head down pose (increased sensitivity)
        head_down_detected = False
        
        if results_pose[0].keypoints is not None:
            for kps in results_pose[0].keypoints.data:
                kps_np = kps.cpu().numpy()
                if check_head_down(kps_np):
                    head_down_detected = True
                    break
        
        # Capture logic: Phone detected OR Head down (Suspected)
        if head_down_detected:
            current_time = time.time()
            if current_time - last_capture_time > CAPTURE_COOLDOWN:
                local_time = get_local_time()
                timestamp_file = local_time.strftime("%Y%m%d_%H%M%S")
                timestamp_display = local_time.strftime("%H:%M:%S")
                date_display = local_time.strftime("%Y-%m-%d")
                
                filename = f"capture_{timestamp_file}.jpg"
                filepath = os.path.join(CAPTURE_FOLDER, filename)
                cv2.imwrite(filepath, annotated_frame)
                
                # Determine alert type
                alert_type = "Phone Usage" if phone_detected else "Suspected Phone Use"
                description = "Person detected using phone" if phone_detected else "Head down pose detected (Suspected phone use)"
                
                # Add context about other objects
                if detected_objects:
                    unique_objs = list(set([o for o in detected_objects if o != "Phone"]))
                    if unique_objs:
                        description += f" near {', '.join(unique_objs)}"
                
                # Save to database
                conn = sqlite3.connect(DATABASE_FILE)
                cursor = conn.cursor()
                cursor.execute('''
                    INSERT INTO phone_alerts (filename, timestamp, date, type, description)
                    VALUES (?, ?, ?, ?, ?)
                ''', (filename, timestamp_display, date_display, alert_type, description))
                conn.commit()
                conn.close()
                
                # Also keep in memory for quick access
                captures.insert(0, {
                    "id": len(captures) + 1,
                    "filename": filename,
                    "timestamp": timestamp_display,
                    "date": date_display,
                    "type": alert_type,
                    "description": description
                })
                if len(captures) > 100:
                    captures.pop()
                last_capture_time = current_time
        
        # Face Recognition - Optimized (run every 3 frames)
        frame_count = getattr(generate_frames, 'frame_count', 0)
        frame_count += 1
        generate_frames.frame_count = frame_count
        
        # Get persistent face results
        face_results = getattr(generate_frames, 'face_results', [])
        
        if len(known_face_encodings) > 0 and frame_count % 3 == 0:
            # Convert BGR to RGB for face_recognition
            rgb_frame = cv2.cvtColor(annotated_frame, cv2.COLOR_BGR2RGB)
            
            # Resize frame more aggressively for faster processing (25% size)
            small_frame = cv2.resize(rgb_frame, (0, 0), fx=0.25, fy=0.25)
            
            # Find faces in frame (use faster HOG model)
            face_locations = face_recognition.face_locations(small_frame, model="hog")
            
            # Prepare new results list
            new_results = []
            
            # Only get encodings if faces found
            if len(face_locations) > 0:
                face_encodings = face_recognition.face_encodings(small_frame, face_locations)
                
                # Process each face
                for (top, right, bottom, left), face_encoding in zip(face_locations, face_encodings):
                    # Scale back up face locations (4x because 25% resize)
                    top *= 4
                    right *= 4
                    bottom *= 4
                    left *= 4
                    
                    # Compare with known faces
                    matches = face_recognition.compare_faces(known_face_encodings, face_encoding, tolerance=0.6)
                    name = "Unknown"
                    
                    # Use the known face with smallest distance
                    face_distances = face_recognition.face_distance(known_face_encodings, face_encoding)
                    if len(face_distances) > 0:
                        best_match_index = np.argmin(face_distances)
                        if matches[best_match_index]:
                            name = known_face_names[best_match_index]
                    
                    new_results.append((top, right, bottom, left, name))
            
            # Update persistent results (clears if no faces found)
            face_results = new_results
            generate_frames.face_results = face_results

        # Draw faces from persistent results (every frame)
        for (top, right, bottom, left, name) in face_results:
            # Draw rectangle around face
            cv2.rectangle(annotated_frame, (left, top), (right, bottom), (0, 255, 0), 2)
            
            # Draw name label
            cv2.rectangle(annotated_frame, (left, bottom - 35), (right, bottom), (0, 255, 0), cv2.FILLED)
            cv2.putText(annotated_frame, name, (left + 6, bottom - 6), 
                       cv2.FONT_HERSHEY_SIMPLEX, 0.8, (255, 255, 255), 2)
        
        current_time = time.time()
        active_person_ids = set()
        
        if results_pose[0].keypoints is not None:
            for idx, kps in enumerate(results_pose[0].keypoints.data):
                kps_np = kps.cpu().numpy()
                status, angle = check_sitting_advanced(kps_np, chairs_boxes)
                
                # Track both Sitting and Standing
                head_x, head_y = int(kps_np[0][0]), int(kps_np[0][1])
                
                if head_x > 0 and head_y > 0:
                    bbox = get_person_bbox(kps_np)
                    
                    if bbox is not None:
                        person_id = match_person_to_session(bbox, current_time, active_person_ids)
                        active_person_ids.add(person_id)
                        
                        if person_id not in person_sessions:
                            person_sessions[person_id] = {
                                "start_time": current_time,
                                "last_seen": current_time,
                                "head_pos": (head_x, head_y),
                                "bbox": bbox,
                                "status": status
                            }
                        else:
                            # Check if status changed
                            old_status = person_sessions[person_id].get("status", "Standing")
                            if old_status != status:
                                # Status changed, reset timer
                                person_sessions[person_id]["start_time"] = current_time
                                person_sessions[person_id]["status"] = status
                            
                            person_sessions[person_id]["last_seen"] = current_time
                            person_sessions[person_id]["head_pos"] = (head_x, head_y)
                            person_sessions[person_id]["bbox"] = bbox
        
        for person_id, session in list(person_sessions.items()):
            time_since_seen = current_time - session["last_seen"]
            status = session.get("status", "Standing")
            
            # Only draw if seen very recently (prevents ghosting/double timers)
            if time_since_seen < 0.5:
                duration = int(current_time - session["start_time"])
                minutes = duration // 60
                seconds = duration % 60
                
                timer_text = f"{minutes}m {seconds}s" if minutes > 0 else f"{seconds}s"
                head_x, head_y = session["head_pos"]
                
                # Draw timer for both Sitting and Standing
                draw_timer_card(annotated_frame, head_x, head_y, timer_text, status.upper())
                
            # Check for session end (for saving to DB)
            if time_since_seen >= SITTING_PERSIST_TIME:
                # Only save SITTING sessions to history/DB
                if status == "Sitting" and person_id not in [s.get('person_id') for s in sitting_history]:
                    duration = int(session["last_seen"] - session["start_time"])
                    if duration >= 5:  # Only save if sat for at least 5 seconds
                        local_time = get_local_time()
                        timestamp_display = local_time.strftime("%H:%M:%S")
                        date_display = local_time.strftime("%Y-%m-%d")
                        
                        # Save to database
                        conn = sqlite3.connect(DATABASE_FILE)
                        cursor = conn.cursor()
                        cursor.execute('''
                            INSERT INTO sitting_sessions (person_id, duration, timestamp, date)
                            VALUES (?, ?, ?, ?)
                        ''', (person_id, duration, timestamp_display, date_display))
                        conn.commit()
                        conn.close()
                        
                        # Also keep in memory
                        sitting_history.insert(0, {
                            "person_id": person_id,
                            "duration": duration,
                            "timestamp": timestamp_display,
                            "date": date_display
                        })
                        if len(sitting_history) > 50:
                            sitting_history.pop()
        
        # Clean up old sessions
        person_sessions = {
            pid: session for pid, session in person_sessions.items()
            if current_time - session["last_seen"] < SITTING_PERSIST_TIME
        }
        
        loop_time = time.time() - loop_start
        if loop_time > 0:
            current_fps = int(1.0 / loop_time)

        # Faster JPEG encoding with lower quality (70 instead of 85)
        ret, buffer = cv2.imencode('.jpg', annotated_frame, [cv2.IMWRITE_JPEG_QUALITY, 70])
        frame_bytes = buffer.tobytes()
        yield (b'--frame\r\n'
               b'Content-Type: image/jpeg\r\n\r\n' + frame_bytes + b'\r\n')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/video_feed')
def video_feed():
    return Response(generate_frames(), mimetype='multipart/x-mixed-replace; boundary=frame')

@app.route('/api/stats')
def get_stats():
    return jsonify({
        "fps": current_fps,
        "capture_count": len(captures),
        "captures": captures[:10],
        "sitting_count": len([s for s in person_sessions.values() if s.get('status') == 'Sitting'])
    })

@app.route('/api/dashboard/stats')
def get_dashboard_stats():
    conn = sqlite3.connect(DATABASE_FILE)
    cursor = conn.cursor()
    
    # Get all phone alerts from database
    cursor.execute('SELECT * FROM phone_alerts ORDER BY created_at DESC LIMIT 50')
    alerts_rows = cursor.fetchall()
    alerts = []
    for row in alerts_rows:
        alerts.append({
            "id": row[0],
            "filename": row[1],
            "timestamp": row[2],
            "date": row[3],
            "type": row[4],
            "description": row[5]
        })
    
    # Get all sitting sessions from database
    cursor.execute('SELECT * FROM sitting_sessions ORDER BY created_at DESC LIMIT 50')
    sitting_rows = cursor.fetchall()
    sitting = []
    for row in sitting_rows:
        sitting.append({
            "id": row[0],
            "person_id": row[1],
            "duration": row[2],
            "timestamp": row[3],
            "date": row[4]
        })
    
    # Calculate statistics
    cursor.execute('SELECT COUNT(*) FROM phone_alerts')
    total_alerts = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM sitting_sessions')
    total_sitting_sessions = cursor.fetchone()[0]
    
    cursor.execute('SELECT AVG(duration) FROM sitting_sessions')
    avg_result = cursor.fetchone()[0]
    avg_sitting_duration = int(avg_result) if avg_result else 0
    
    # Today's stats
    today = get_local_time().strftime("%Y-%m-%d")
    cursor.execute('SELECT COUNT(*) FROM phone_alerts WHERE date = ?', (today,))
    today_alerts = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM sitting_sessions WHERE date = ?', (today,))
    today_sitting = cursor.fetchone()[0]
    
    conn.close()
    
    return jsonify({
        "total_alerts": total_alerts,
        "total_sitting_sessions": total_sitting_sessions,
        "avg_sitting_duration": avg_sitting_duration,
        "today_alerts": today_alerts,
        "today_sitting": today_sitting,
        "current_sitting": len([s for s in person_sessions.values() if s.get('status') == 'Sitting']),
        "captures": alerts,
        "sitting_history": sitting
    })

@app.route('/api/export/excel')
def export_excel():
    conn = sqlite3.connect(DATABASE_FILE)
    cursor = conn.cursor()
    
    # Create workbook
    wb = Workbook()
    
    # Phone Alerts Sheet
    ws1 = wb.active
    ws1.title = "Phone Alerts"
    
    # Headers
    headers1 = ["ID", "Date", "Time", "Type", "Description", "Filename"]
    ws1.append(headers1)
    
    # Style headers
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws1[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Get data
    cursor.execute('SELECT id, date, timestamp, type, description, filename FROM phone_alerts ORDER BY created_at DESC')
    for row in cursor.fetchall():
        ws1.append(row)
    
    # Sitting Sessions Sheet
    ws2 = wb.create_sheet("Sitting Sessions")
    headers2 = ["ID", "Person ID", "Date", "Time", "Duration (seconds)"]
    ws2.append(headers2)
    
    # Style headers
    for cell in ws2[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Get data
    cursor.execute('SELECT id, person_id, date, timestamp, duration FROM sitting_sessions ORDER BY created_at DESC')
    for row in cursor.fetchall():
        ws2.append(row)
    
    conn.close()
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    filename = f"visionguard_export_{get_local_time().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

@app.route('/api/settings', methods=['POST'])
def update_settings():
    data = request.json
    if 'fps_limit' in data:
        settings['fps_limit'] = int(data['fps_limit'])
    if 'mode' in data:
        settings['mode'] = data['mode']
        settings['conf_threshold'] = 0.35 if data['mode'] == 'accurate' else 0.25
    if 'low_light_mode' in data:
        settings['low_light_mode'] = bool(data['low_light_mode'])
    return jsonify({"status": "success", "settings": settings})

@app.route('/api/reset_db', methods=['POST'])
def reset_db():
    try:
        # Clear database tables
        conn = sqlite3.connect(DATABASE_FILE)
        cursor = conn.cursor()
        cursor.execute('DELETE FROM phone_alerts')
        cursor.execute('DELETE FROM sitting_sessions')
        conn.commit()
        conn.close()
        
        # Clear captures folder
        for filename in os.listdir(CAPTURE_FOLDER):
            file_path = os.path.join(CAPTURE_FOLDER, filename)
            try:
                if os.path.isfile(file_path) or os.path.islink(file_path):
                    os.unlink(file_path)
            except Exception as e:
                print(f'Failed to delete {file_path}. Reason: {e}')
                
        # Clear in-memory lists
        global captures, person_sessions, sitting_history
        captures = []
        person_sessions = {}
        sitting_history = []
        
        return jsonify({"status": "success", "message": "Database and captures reset successfully"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500

@app.route('/captures/<path:filename>')
def serve_capture(filename):
    return send_from_directory(CAPTURE_FOLDER, filename)

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=2123, debug=False, threaded=True)
